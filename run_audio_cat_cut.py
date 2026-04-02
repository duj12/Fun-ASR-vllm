#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Audio concat and split tool

Stage 1: Group by duration -> pad silence -> concat to ~12h audio files
Stage 2: Align recorded audio with original (cross-correlation offset detection)
Stage 3: Split aligned audio by original segment duration, restore text
Stage 4: ASR evaluation on split segments, output WER to Excel

Merged: align_split_asr = Stage 2 + 3 + 4（1ch 与 4ch 各输出切分目录与 WER Excel）
Stage 5: 仅按单通道 WER 筛选，四通道按同行下标同步；结果写入 segments/ 与 segments_4ch/；完成后删除 aligned/
"""

import os
import sys
import shutil
import argparse
import logging
import tempfile
import numpy as np
import soundfile as sf
import librosa
from pathlib import Path
from typing import List, Tuple, Dict, Optional, Any

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)

SAMPLE_RATE = 16000
GROUP_CFGS = [
    {"name": "10s", "max_dur": 10.0, "pad_to": 10.0},
    {"name": "20s", "max_dur": 20.0, "pad_to": 20.0},
    {"name": "30s", "max_dur": 30.0, "pad_to": 30.0},
]
HOURS_PER_FILE = 12
# 4ch float32->int16 写盘时分块帧数，避免与 aligned_4ch 同尺寸临时数组导致峰值内存翻倍
WRITE_4CH_PCM_CHUNK_FRAMES = 400_000

# 各子命令中需要规范化的路径参数名（兼容 Windows 绝对路径、反斜杠、外层引号）
_PATH_ARG_NAMES = (
    "wav_scp",
    "text_tn",
    "text_itn",
    "wav2dur",
    "output_dir",
    "concat_wav",
    "recorded_1ch",
    "recorded_4ch",
    "output_1ch",
    "output_4ch",
    "aligned_wav",
    "concat_tn_txt",
    "concat_itn_txt",
    "aligned_wav_4ch",
    "output_dir_4ch",
    "segments_dir",
    "text_file",
    "output_excel",
    "work_dir",
    "output_excel_4ch",
    "input_dir",
    "excel_ch1",
    "excel_ch4",
)


def normalize_fs_path(path: str) -> str:
    """统一文件系统路径，兼容 Windows 绝对路径、混合斜杠、复制时带上的引号。"""
    if not isinstance(path, str):
        return path
    p = path.strip().strip('"').strip("'")
    if not p:
        return p
    p = os.path.expanduser(p)
    return os.path.normpath(p)


def normalize_cli_paths(args: argparse.Namespace) -> None:
    for name in _PATH_ARG_NAMES:
        if not hasattr(args, name):
            continue
        v = getattr(args, name)
        if v is None or not isinstance(v, str):
            continue
        setattr(args, name, normalize_fs_path(v))


def resolve_align_split_input_dir(input_dir: str) -> Dict[str, str]:
    """从 input_dir 解析 align_split_asr 所需的 5 个文件路径。

    - 唯一 ``.wav``：拼接参考音频
    - 文件名以 ``ch1.pcm`` 结尾：单通道 PCM
    - 文件名以 ``ch4.pcm`` 结尾：四通道 PCM
    - 文件名以 ``_tn.txt`` 结尾：concat tn 文本
    - 文件名以 ``_itn.txt`` 结尾：concat itn 文本
    """
    d = normalize_fs_path(input_dir)
    if not os.path.isdir(d):
        raise ValueError(f"input_dir 不是目录: {d}")
    names = [n for n in os.listdir(d) if not n.startswith(".")]

    def pick_one(suffix: str, desc: str) -> str:
        cands = [n for n in names if n.endswith(suffix)]
        if len(cands) != 1:
            raise ValueError(
                f"input_dir 中期望恰好 1 个以 {suffix!r} 结尾的文件（{desc}），实际 {len(cands)}: {cands}"
            )
        return os.path.join(d, cands[0])

    wavs = [n for n in names if n.lower().endswith(".wav")]
    if len(wavs) != 1:
        raise ValueError(
            f"input_dir 中期望恰好 1 个 .wav（拼接参考音频），实际 {len(wavs)}: {wavs}"
        )
    return {
        "concat_wav": os.path.join(d, wavs[0]),
        "recorded_1ch": pick_one("ch1.pcm", "单通道 PCM"),
        "recorded_4ch": pick_one("ch4.pcm", "四通道 PCM"),
        "concat_tn_txt": pick_one("_tn.txt", "tn 文本"),
        "concat_itn_txt": pick_one("_itn.txt", "itn 文本"),
    }


def read_kv_file(path: str) -> Dict[str, str]:
    data = {}
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = line.split(maxsplit=1)
            if len(parts) != 2:
                logger.warning(f"skip bad line: {line}")
                continue
            data[parts[0]] = parts[1]
    return data


def load_audio_mono(path: str, target_sr: int = SAMPLE_RATE) -> np.ndarray:
    #audio, sr = sf.read(path, dtype="float32")
    audio, sr = librosa.load(path, sr=target_sr, mono=True)
    if audio.ndim > 1:
        audio = audio.mean(axis=1)
    if sr != target_sr:
        raise ValueError(f"SR mismatch: {path} is {sr}Hz, expected {target_sr}Hz")
    return audio


def pad_silence(audio: np.ndarray, target_samples: int) -> np.ndarray:
    if len(audio) >= target_samples:
        return audio[:target_samples]
    return np.pad(audio, (0, target_samples - len(audio)), mode="constant")


def pad_silence_mc(audio: np.ndarray, target_samples: int) -> np.ndarray:
    """Pad silence for multi-channel audio (N, C)."""
    if audio.ndim == 1:
        return pad_silence(audio, target_samples)
    if len(audio) >= target_samples:
        return audio[:target_samples]
    return np.pad(audio, ((0, target_samples - len(audio)), (0, 0)), mode="constant")


def load_pcm(path: str, channels: int = 1, dtype=np.int16,
             sr: int = SAMPLE_RATE) -> np.ndarray:
    """Load raw PCM file -> float32. Returns (N,) for mono or (N, C) for multi-channel."""
    raw = np.fromfile(path, dtype=dtype)
    if channels > 1:
        remainder = len(raw) % channels
        if remainder != 0:
            logger.warning(
                f"PCM sample count {len(raw)} not divisible by {channels}, "
                f"truncating last {remainder} samples"
            )
            raw = raw[:len(raw) - remainder]
        raw = raw.reshape(-1, channels)
    audio = raw.astype(np.float32) / np.iinfo(dtype).max
    logger.info(
        f"loaded PCM {path}: shape={audio.shape}, "
        f"dur={len(audio)/sr:.2f}s, range=[{audio.min():.4f}, {audio.max():.4f}]"
    )
    return audio


def write_text_list(path: str, items: List[Tuple[str, str]]):
    with open(path, "w", encoding="utf-8") as f:
        for utt_id, text in items:
            f.write(f"{utt_id}\t{text}\n")


def stage1_concat(wav_scp, text_tn_file, text_itn_file, wav2dur_file,
                  output_dir, sample_rate=SAMPLE_RATE):
    wav_dict = read_kv_file(wav_scp)
    tn_dict = read_kv_file(text_tn_file)
    itn_dict = read_kv_file(text_itn_file)
    dur_dict = read_kv_file(wav2dur_file)
    os.makedirs(output_dir, exist_ok=True)

    groups: Dict[str, List[str]] = {cfg["name"]: [] for cfg in GROUP_CFGS}
    for utt_id, dur_str in dur_dict.items():
        dur = float(dur_str)
        if utt_id not in wav_dict:
            logger.warning(f"wav.scp missing {utt_id}, skip")
            continue
        assigned = False
        for cfg in GROUP_CFGS:
            if dur <= cfg["max_dur"]:
                groups[cfg["name"]].append(utt_id)
                assigned = True
                break
        if not assigned:
            logger.info(f"dur {dur:.2f}s > 30s, skip {utt_id}")

    samples_per_file = int(HOURS_PER_FILE * 3600 * sample_rate)

    for cfg in GROUP_CFGS:
        gname = cfg["name"]
        pad_samples = int(cfg["pad_to"] * sample_rate)
        utt_ids = groups[gname]
        if not utt_ids:
            logger.info(f"group {gname}: no data, skip")
            continue
        logger.info(f"group {gname}: {len(utt_ids)} utts")

        buf = np.array([], dtype=np.float32)
        tn_buf: List[Tuple[str, str]] = []
        itn_buf: List[Tuple[str, str]] = []
        fidx = 1

        for i, utt_id in enumerate(utt_ids):
            audio = load_audio_mono(wav_dict[utt_id], sample_rate)
            audio = pad_silence(audio, pad_samples)
            buf = np.concatenate([buf, audio])
            tn_buf.append((utt_id, tn_dict.get(utt_id, "")))
            itn_buf.append((utt_id, itn_dict.get(utt_id, "")))

            if len(buf) >= samples_per_file or i == len(utt_ids) - 1:
                out_name = f"{gname}_{fidx:02d}"
                sf.write(os.path.join(output_dir, f"{out_name}.wav"), buf, sample_rate)
                write_text_list(os.path.join(output_dir, f"{out_name}_tn.txt"), tn_buf)
                write_text_list(os.path.join(output_dir, f"{out_name}_itn.txt"), itn_buf)
                logger.info(
                    f"  {out_name}.wav  dur={len(buf)/sample_rate/3600:.2f}h  n={len(tn_buf)}"
                )
                buf = np.array([], dtype=np.float32)
                tn_buf = []
                itn_buf = []
                fidx += 1


def _find_speech_onset(audio, sr, frame_ms=10,
                       threshold_ratio=0.05, min_consecutive=3):
    """Return sample index where speech begins (energy-based)."""
    frame_len = int(sr * frame_ms / 1000)
    n_frames = len(audio) // frame_len
    if n_frames == 0:
        return 0
    rms = np.array([
        np.sqrt(np.mean(audio[i * frame_len:(i + 1) * frame_len] ** 2))
        for i in range(n_frames)
    ])
    thresh = rms.max() * threshold_ratio
    streak = 0
    for i, r in enumerate(rms):
        if r > thresh:
            streak += 1
            if streak >= min_consecutive:
                return max(0, (i - min_consecutive + 1) * frame_len)
        else:
            streak = 0
    return 0


def _find_offset_xcorr(ref, rec, search_range_sec=60.0, sr=SAMPLE_RATE):
    """Normalized cross-correlation with speech-anchored template.

    Returns *offset* such that ``rec[offset:]`` aligns with ``ref[0:]``.

    Key fixes over the previous version:
    * Template starts at the first speech onset in *ref* (skips leading silence
      and avoids using a mostly-silent template).
    * Normalised cross-correlation (NCC) so correlation magnitude is
      independent of local energy.
    * Correct k -> position mapping: ``conv[k]`` from the FFT convolution
      corresponds to the template placed at position ``k - (tpl_len - 1)``,
      not at ``k``.
    """
    from numpy.fft import fft, ifft

    ref_onset = _find_speech_onset(ref, sr)
    tpl_dur = min(5.0, (len(ref) - ref_onset) / sr)
    tpl_len = int(tpl_dur * sr)
    if tpl_len == 0:
        return 0
    tpl = ref[ref_onset:ref_onset + tpl_len].astype(np.float64)

    search_samples = int(search_range_sec * sr) + tpl_len
    region = rec[:min(search_samples, len(rec))].astype(np.float64)
    if len(region) < tpl_len:
        return 0

    n = len(region)

    # --- FFT cross-correlation ---
    t = np.zeros(n, dtype=np.float64)
    t[:tpl_len] = tpl[::-1]
    xcorr = np.real(ifft(fft(region) * fft(t)))

    # --- normalise ---
    tpl_energy = np.sum(tpl ** 2)
    cum = np.cumsum(region ** 2)
    valid_len = n - tpl_len + 1          # number of valid positions
    win_energy = np.empty(valid_len)
    win_energy[0] = cum[tpl_len - 1]
    if valid_len > 1:
        win_energy[1:] = cum[tpl_len:] - cum[:valid_len - 1]
    denom = np.sqrt(tpl_energy * win_energy + 1e-12)

    # conv[k] <-> template at position (k - tpl_len + 1)
    # valid positions: 0 .. valid_len-1  <->  k = tpl_len-1 .. n-1
    ncc = xcorr[tpl_len - 1:tpl_len - 1 + valid_len] / denom

    best_pos = int(np.argmax(ncc))       # position in *region*
    # ref[ref_onset] matches rec[best_pos]  =>  ref[0] matches rec[best_pos - ref_onset]
    offset = best_pos - ref_onset

    logger.info(
        f"xcorr: ref_onset={ref_onset}({ref_onset / sr:.3f}s) "
        f"match_pos={best_pos}({best_pos / sr:.3f}s) "
        f"ncc={ncc[best_pos]:.4f} offset={offset}({offset / sr:.3f}s)"
    )
    return max(0, offset)


def stage2_align(concat_wav, recorded_1ch_pcm, recorded_4ch_pcm,
                 output_1ch_wav, output_4ch_wav,
                 search_range_sec=60.0, sample_rate=SAMPLE_RATE):
    ref = load_audio_mono(concat_wav, sample_rate)
    rec_1ch = load_pcm(recorded_1ch_pcm, channels=1, sr=sample_rate)
    rec_4ch = load_pcm(recorded_4ch_pcm, channels=4, sr=sample_rate)

    ref_len = len(ref)

    # ---- align 1ch independently ----
    logger.info("=== aligning 1ch ===")
    offset_1ch = _find_offset_xcorr(ref, rec_1ch, search_range_sec, sample_rate)

    aligned_1ch = rec_1ch[offset_1ch:offset_1ch + ref_len]
    if len(aligned_1ch) < ref_len:
        aligned_1ch = pad_silence(aligned_1ch, ref_len)
        logger.warning("1ch recorded audio too short, padded silence at tail")

    # ---- align 4ch independently (correlate on channel-mean) ----
    logger.info("=== aligning 4ch ===")
    rec_4ch_mono = rec_4ch.mean(axis=1)
    offset_4ch = _find_offset_xcorr(ref, rec_4ch_mono, search_range_sec, sample_rate)

    aligned_4ch = rec_4ch[offset_4ch:offset_4ch + ref_len]
    if len(aligned_4ch) < ref_len:
        aligned_4ch = pad_silence_mc(aligned_4ch, ref_len)
        logger.warning("4ch recorded audio too short, padded silence at tail")

    if offset_1ch != offset_4ch:
        logger.warning(
            f"1ch/4ch offsets differ: "
            f"1ch={offset_1ch}({offset_1ch / sample_rate:.3f}s)  "
            f"4ch={offset_4ch}({offset_4ch / sample_rate:.3f}s)"
        )

    os.makedirs(os.path.dirname(output_1ch_wav) or ".", exist_ok=True)
    os.makedirs(os.path.dirname(output_4ch_wav) or ".", exist_ok=True)

    sf.write(output_1ch_wav, aligned_1ch, sample_rate)
    # 4ch：分块 float32->int16 并流式写入，避免整段 (aligned_4ch * scale) 再占一份 ~10GB 内存
    n_frames, n_ch = aligned_4ch.shape[0], aligned_4ch.shape[1]
    data_bytes = n_frames * n_ch * np.dtype(np.int16).itemsize
    fmt = "RF64" if data_bytes >= 4 * 1024 ** 3 else "WAV"
    logger.info(
        f"4ch data size: {data_bytes / 1024**3:.2f}GB, using {fmt} format, "
        f"write chunk={WRITE_4CH_PCM_CHUNK_FRAMES} frames"
    )
    chunk = WRITE_4CH_PCM_CHUNK_FRAMES
    with sf.SoundFile(
        output_4ch_wav, "w", samplerate=sample_rate, channels=n_ch,
        subtype="PCM_16", format=fmt,
    ) as f:
        for start in range(0, n_frames, chunk):
            end = min(start + chunk, n_frames)
            n = end - start
            sl = aligned_4ch[start:end]
            work = np.empty((n, n_ch), dtype=np.float32, order="C")
            np.multiply(sl, 32767.0, out=work)
            np.clip(work, -32768, 32767, out=work)
            f.write(work.astype(np.int16, copy=False))
    logger.info(
        f"aligned 1ch -> {output_1ch_wav}  "
        f"offset={offset_1ch}({offset_1ch / sample_rate:.3f}s)  "
        f"dur={len(aligned_1ch) / sample_rate / 3600:.2f}h"
    )
    logger.info(
        f"aligned 4ch -> {output_4ch_wav}  "
        f"offset={offset_4ch}({offset_4ch / sample_rate:.3f}s)  "
        f"shape={aligned_4ch.shape}  "
        f"dur={len(aligned_4ch) / sample_rate / 3600:.2f}h"
    )


def _read_text_items(path: str) -> List[Tuple[str, str]]:
    items = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = line.split("\t", 1)
            items.append((parts[0], parts[1] if len(parts) > 1 else ""))
    return items


def stage3_split(aligned_wav, concat_tn_txt, concat_itn_txt, output_dir,
                 segment_sec=10.0, sample_rate=SAMPLE_RATE,
                 aligned_wav_4ch=None, output_dir_4ch=None):
    audio = load_audio_mono(aligned_wav, sample_rate)
    seg_samples = int(segment_sec * sample_rate)

    tn_items = _read_text_items(concat_tn_txt)
    itn_items = _read_text_items(concat_itn_txt)
    if len(tn_items) != len(itn_items):
        logger.warning(
            f"text_tn ({len(tn_items)}) and text_itn ({len(itn_items)}) "
            f"have different line counts, using min"
        )
    n_segs = min(len(tn_items), len(itn_items))

    base_name = Path(aligned_wav).stem
    os.makedirs(output_dir, exist_ok=True)
    new_tn: List[Tuple[str, str]] = []
    new_itn: List[Tuple[str, str]] = []

    for idx in range(n_segs):
        start = idx * seg_samples
        end = start + seg_samples
        seg = audio[start:end]
        if len(seg) < seg_samples:
            seg = pad_silence(seg, seg_samples)
            logger.warning(f"seg {idx+1} too short, padded")
        seg_name = f"{base_name}_{idx + 1:04d}"
        sf.write(os.path.join(output_dir, f"{seg_name}.wav"), seg, sample_rate)
        new_tn.append((seg_name, tn_items[idx][1]))
        new_itn.append((seg_name, itn_items[idx][1]))

    write_text_list(os.path.join(output_dir, f"{base_name}_text_tn.txt"), new_tn)
    write_text_list(os.path.join(output_dir, f"{base_name}_text_itn.txt"), new_itn)
    logger.info(f"split done (1ch): {len(new_tn)} segs -> {output_dir}")

    if aligned_wav_4ch is not None and output_dir_4ch is not None:
        audio_4ch, sr = sf.read(aligned_wav_4ch, dtype="float32")
        if audio_4ch.ndim == 1:
            audio_4ch = audio_4ch[:, np.newaxis]
        if sr != sample_rate:
            raise ValueError(
                f"4ch WAV sample rate {sr} != expected {sample_rate}"
            )
        base_name_4ch = Path(aligned_wav_4ch).stem
        os.makedirs(output_dir_4ch, exist_ok=True)
        new_tn_4ch: List[Tuple[str, str]] = []
        new_itn_4ch: List[Tuple[str, str]] = []

        for idx in range(n_segs):
            start = idx * seg_samples
            end = start + seg_samples
            seg_4ch = audio_4ch[start:end]
            if len(seg_4ch) < seg_samples:
                seg_4ch = pad_silence_mc(seg_4ch, seg_samples)
                logger.warning(f"4ch seg {idx+1} too short, padded")
            seg_name_4ch = f"{base_name_4ch}_{idx + 1:04d}"
            sf.write(
                os.path.join(output_dir_4ch, f"{seg_name_4ch}.wav"),
                seg_4ch,
                sample_rate,
            )
            new_tn_4ch.append((seg_name_4ch, tn_items[idx][1]))
            new_itn_4ch.append((seg_name_4ch, itn_items[idx][1]))

        write_text_list(
            os.path.join(output_dir_4ch, f"{base_name_4ch}_text_tn.txt"),
            new_tn_4ch,
        )
        write_text_list(
            os.path.join(output_dir_4ch, f"{base_name_4ch}_text_itn.txt"),
            new_itn_4ch,
        )
        logger.info(f"split done (4ch): {len(new_tn_4ch)} segs -> {output_dir_4ch}")


def _mono_wav_for_asr(
    audio_path: str, multichannel_downmix: bool
) -> Tuple[str, bool]:
    """若需对多声道做 ASR，先下混为单声道临时 WAV。

    返回 (供 transcribe 使用的路径, 是否为需删除的临时文件)。
    Qwen3-ASR 等模型通常只接受单声道输入。
    """
    if not multichannel_downmix:
        return audio_path, False
    try:
        info = sf.info(audio_path)
    except OSError as e:
        logger.warning(f"无法读取音频信息 {audio_path}: {e}")
        return audio_path, False
    if info.channels <= 1:
        return audio_path, False
    data, sr = sf.read(audio_path, dtype="float32")
    if data.ndim == 1:
        return audio_path, False
    mono = np.mean(data, axis=1)
    fd, tmp_path = tempfile.mkstemp(suffix=".wav", prefix="asr_mono_")
    os.close(fd)
    sf.write(tmp_path, mono, sr)
    return tmp_path, True


def stage4_asr_eval(segments_dir: str, text_file: str, output_excel: str,
                   asr_model: str = "./Qwen/Qwen3-ASR-1.7B", batch_size: int = 16,
                   device: str = "cuda:0",
                   multichannel_downmix: bool = False,
                   preloaded_asr_model: Optional[Any] = None) -> Optional[Any]:
    """Stage 4: 使用 Qwen3-ASR 对切分音频做识别（支持 batch），用 compute_wer_line 算 WER，写入 Excel（wav_name, text, asr, wer）。

    multichannel_downmix: 为 True 时对多声道 WAV 先按通道平均下混为单声道再识别（4 通道切分片段需要）。
    preloaded_asr_model: 若传入则不再加载模型（合并流程中 1ch/4ch 共用一个模型）。
    返回本次推理使用的 ASR 模型对象（便于调用方复用）；未加载或未跑完则返回 None。
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("stage4 需要 openpyxl，请执行: pip install openpyxl")

    try:
        from qwen_asr import Qwen3ASRModel
    except ImportError:
        raise ImportError("stage4 使用 Qwen3-ASR 需安装 qwen_asr，请按项目说明安装")

    from asr_check import determine_lang
    from compute_wer_line import compute_wer_line

    items = _read_text_items(text_file)
    if not items:
        logger.warning("text_file 为空，未生成任何结果")
        return None

    lang_map = {"zh": "Chinese", "en": "English"}
    # 预扫：有效条目 (items 下标, wav_name, text, audio_path, language)；缺失的先填行
    rows = [None] * len(items)
    valid_list = []
    for i, (wav_name, text) in enumerate(items):
        audio_path = os.path.join(segments_dir, f"{wav_name}.wav")
        if not os.path.isfile(audio_path):
            logger.warning(f"跳过缺失音频: {audio_path}")
            rows[i] = {"wav_name": wav_name, "text": text, "asr": "", "wer": ""}
            continue
        lang_code = determine_lang(text)
        language = lang_map.get(lang_code, "Chinese")
        valid_list.append((i, wav_name, text, audio_path, language))

    if not valid_list:
        logger.warning("没有可用的音频文件，未生成结果")
        return None

    # 使用单卡避免 device_map="auto" 把模型拆到多卡导致 tensor 设备不一致
    if preloaded_asr_model is not None:
        asr_model_obj = preloaded_asr_model
        logger.info("复用已加载的 ASR 模型进行推理")
    else:
        import torch
        logger.info(f"加载 ASR 模型: {asr_model}，batch_size={batch_size}，device={device}")
        asr_model_obj = Qwen3ASRModel.from_pretrained(
            asr_model,
            device_map=device,
            dtype=torch.bfloat16,
            max_inference_batch_size=max(batch_size, 32),
            max_new_tokens=256,
        )

    # 按 batch 推理，每推理完一个 batch 立即算该 batch 的 WER 并填入 rows
    n_valid = len(valid_list)
    for start in range(0, n_valid, batch_size):
        batch = valid_list[start : start + batch_size]
        paths_raw = [x[3] for x in batch]
        langs = [x[4] for x in batch]
        temp_paths: List[str] = []
        paths: List[str] = []
        for p in paths_raw:
            use_p, is_temp = _mono_wav_for_asr(p, multichannel_downmix)
            paths.append(use_p)
            if is_temp:
                temp_paths.append(use_p)
        try:
            batch_results = asr_model_obj.transcribe(
                audio=paths,
                language=langs,
                return_time_stamps=False,
            )
            for k, (idx, wav_name, text, _path, _lang) in enumerate(batch):
                asr = batch_results[k].text if k < len(batch_results) and batch_results[k] else ""
                try:
                    wer_result = compute_wer_line(text, asr, tochar=True, verbose=1)
                    wer = wer_result["stats"]["wer"]
                except Exception:
                    wer = float("nan")
                rows[idx] = {"wav_name": wav_name, "text": text, "asr": asr, "wer": wer}
        except Exception as e:
            logger.warning(f"batch {start}-{start + len(batch)} 推理失败: {e}")
            for idx, wav_name, text, _path, _lang in batch:
                rows[idx] = {"wav_name": wav_name, "text": text, "asr": "", "wer": float("nan")}
        finally:
            for tp in temp_paths:
                try:
                    os.unlink(tp)
                except OSError:
                    pass
        if (start + len(batch)) % (batch_size * 10) == 0 or start + len(batch) == n_valid:
            logger.info(f"已推理 {min(start + len(batch), n_valid)}/{n_valid} 条")

    rows = [r for r in rows if r is not None]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ASR"
    ws.append(["wav_name", "text", "asr", "wer"])
    for r in rows:
        ws.append([r["wav_name"], r["text"], r["asr"], r["wer"]])

    os.makedirs(os.path.dirname(output_excel) or ".", exist_ok=True)
    wb.save(output_excel)
    logger.info(f"ASR 结果已写入: {output_excel} 共 {len(rows)} 条")
    return asr_model_obj


def _parse_wer_range_spec(spec: str) -> Tuple[int, int, float]:
    """解析 ``start:end:threshold``，序号区间为闭区间 [start, end]（与 Excel 数据行 0 起始下标一致）。"""
    parts = spec.strip().split(":")
    if len(parts) != 3:
        raise ValueError(f"范围格式应为 start:end:threshold，收到: {spec!r}")
    a, b, t = int(parts[0].strip()), int(parts[1].strip()), float(parts[2].strip())
    if a > b:
        a, b = b, a
    return a, b, t


def _effective_wer_threshold(
    row_index: int,
    ranges: List[Tuple[int, int, float]],
    global_max_wer: float,
) -> float:
    """行 ``row_index`` 上允许的最大 wer：全局上限与各包含该行的区间阈值取最小。"""
    applicable = [thr for lo, hi, thr in ranges if lo <= row_index <= hi]
    if applicable:
        return min(global_max_wer, min(applicable))
    return global_max_wer


def _cell_to_wer_float(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _read_excel_rows(path: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    try:
        import openpyxl
    except ImportError:
        raise ImportError("Stage 5 需要 openpyxl，请执行: pip install openpyxl")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        wb.close()
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    data = []
    for row in rows_iter:
        d = {}
        for i, h in enumerate(headers):
            if h:
                d[h] = row[i] if i < len(row) else None
        data.append(d)
    wb.close()
    return headers, data


def _compute_kept_row_indices_ch1(
    data: List[Dict[str, Any]],
    ranges: List[Tuple[int, int, float]],
    global_max_wer: float,
) -> List[int]:
    """仅按单通道 WER 规则得到保留行的下标（0 起始）；wer 无效的行不保留。"""
    kept_idx: List[int] = []
    for i, row in enumerate(data):
        lower = {k.lower(): v for k, v in row.items()}
        wer_raw = lower.get("wer")
        w = _cell_to_wer_float(wer_raw)
        if w is None:
            continue
        thr = _effective_wer_threshold(i, ranges, global_max_wer)
        if w <= thr:
            kept_idx.append(i)
    return kept_idx


def _filter_excel_rows(
    data: List[Dict[str, Any]],
    ranges: List[Tuple[int, int, float]],
    global_max_wer: float,
) -> List[Dict[str, Any]]:
    """按行下标（0 起始）与 wer 阈值保留行（单通道逻辑）。"""
    idx = _compute_kept_row_indices_ch1(data, ranges, global_max_wer)
    return [data[i] for i in idx]


def _write_filtered_excel(
    out_path: str,
    rows: List[Dict[str, Any]],
    source_headers: List[str],
) -> None:
    try:
        import openpyxl
    except ImportError:
        raise ImportError("Stage 5 需要 openpyxl，请执行: pip install openpyxl")

    # 列顺序：wav_name / text 重命名，其余保持原列名
    key_order = []
    for h in source_headers:
        if not h:
            continue
        hl = h.lower()
        if hl == "wav_name":
            key_order.append((h, "音频名称"))
        elif hl == "text":
            key_order.append((h, "标注后文本"))
        else:
            key_order.append((h, h))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ASR"
    ws.append([disp for _src, disp in key_order])
    for row in rows:
        line = []
        for src_key, _disp in key_order:
            line.append(row.get(src_key, ""))
        ws.append(line)

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)


def stage5_filter_wer(
    output_dir: str,
    range_specs: List[str],
    global_max_wer: float = 1.0,
    excel_ch1: Optional[str] = None,
    excel_ch4: Optional[str] = None,
):
    """Stage 5: **仅对单通道** WER Excel 做区间筛选；四通道按相同行下标同步子集，保证 1ch/4ch 条数一致。

    输出：单通道 -> ``output_dir/segments/asr_wer_ch1_filter.xlsx``，四通道 -> ``output_dir/segments_4ch/asr_wer_ch4_filter.xlsx``（与 ch1 保留行一一对应）。
    处理结束后若存在 ``output_dir/aligned/``（第二步对齐产物），则整目录删除以节省磁盘。
    数据行下标为 **0 起始**（与 Excel 中首条数据行对应下标 0），与 Stage3/4 切分顺序一致。
    """
    d = normalize_fs_path(output_dir)
    if not os.path.isdir(d):
        raise ValueError(f"output_dir 不是目录: {d}")

    ranges = [_parse_wer_range_spec(s) for s in range_specs]
    if excel_ch1 is None:
        excel_ch1 = os.path.join(d, "asr_wer_ch1.xlsx")
    if excel_ch4 is None:
        excel_ch4 = os.path.join(d, "asr_wer_ch4.xlsx")

    segments_dir = os.path.join(d, "segments")
    segments_4ch_dir = os.path.join(d, "segments_4ch")

    if not os.path.isfile(excel_ch1):
        raise ValueError(f"缺少单通道 WER Excel（筛选依据）: {excel_ch1}")

    headers1, data1 = _read_excel_rows(excel_ch1)
    n1 = len(data1)
    kept_idx = _compute_kept_row_indices_ch1(data1, ranges, global_max_wer)
    kept_ch1 = [data1[i] for i in kept_idx]

    p1 = Path(excel_ch1)
    os.makedirs(segments_dir, exist_ok=True)
    out_ch1 = os.path.join(segments_dir, f"{p1.stem}_filter{p1.suffix}")
    _write_filtered_excel(out_ch1, kept_ch1, headers1)
    logger.info(
        "ch1: %s -> %s 保留 %d/%d 行 (global_max_wer=%s, ranges=%s)",
        excel_ch1,
        out_ch1,
        len(kept_ch1),
        n1,
        global_max_wer,
        ranges,
    )

    if os.path.isfile(excel_ch4):
        headers4, data4 = _read_excel_rows(excel_ch4)
        n4 = len(data4)
        if n4 != n1:
            logger.warning(
                "4ch 行数 (%d) 与 ch1 (%d) 不一致，按 ch1 下标同步；缺失下标将跳过",
                n4,
                n1,
            )
        kept_4ch: List[Dict[str, Any]] = []
        for i in kept_idx:
            if i < n4:
                kept_4ch.append(data4[i])
            else:
                logger.warning("4ch 缺少下标 %d，未写入对应行", i)
        p4 = Path(excel_ch4)
        os.makedirs(segments_4ch_dir, exist_ok=True)
        out_path_4 = os.path.join(segments_4ch_dir, f"{p4.stem}_filter{p4.suffix}")
        _write_filtered_excel(out_path_4, kept_4ch, headers4)
        logger.info(
            "ch4: %s -> %s 同步 %d 行（与 ch1 筛选一致，未再按 4ch 的 wer 筛选）",
            excel_ch4,
            out_path_4,
            len(kept_4ch),
        )
    else:
        logger.warning("未找到 4ch Excel，跳过同步: %s", excel_ch4)

    aligned_dir = os.path.join(d, "aligned")
    if os.path.isdir(aligned_dir):
        try:
            shutil.rmtree(aligned_dir)
            logger.info("已删除对齐目录以节省空间: %s", aligned_dir)
        except OSError as e:
            logger.warning("删除对齐目录失败 %s: %s", aligned_dir, e)


def run_align_split_asr(
    concat_wav: str,
    recorded_1ch: str,
    recorded_4ch: str,
    concat_tn_txt: str,
    concat_itn_txt: str,
    segment_sec: float,
    work_dir: str,
    output_excel: Optional[str] = None,
    output_excel_4ch: Optional[str] = None,
    search_range_sec: float = 30.0,
    sample_rate: int = SAMPLE_RATE,
    text_type: str = "tn",
    asr_model: str = "./Qwen/Qwen3-ASR-1.7B",
    batch_size: int = 16,
    device: str = "cuda:0",
):
    """合并执行：对齐 -> 切分 -> ASR 转写并输出 WER Excel（1ch 与 4ch 各一份）。

    中间产物：work_dir/aligned/ 下为对齐后的 1ch/4ch WAV；
             work_dir/segments/、work_dir/segments_4ch/ 下为切分片段与文本。
    4 通道片段为多声道 WAV，ASR 前会下混为单声道再识别。
    output_excel / output_excel_4ch 为 None 时默认为 work_dir/asr_wer_ch1.xlsx 与 work_dir/asr_wer_ch4.xlsx。
    """
    work_dir = normalize_fs_path(work_dir)
    if output_excel is None:
        output_excel = os.path.join(work_dir, "asr_wer_ch1.xlsx")
    if output_excel_4ch is None:
        output_excel_4ch = os.path.join(work_dir, "asr_wer_ch4.xlsx")

    concat_stem = Path(concat_wav).stem
    aligned_dir = os.path.join(work_dir, "aligned")
    segments_dir = os.path.join(work_dir, "segments")
    segments_dir_4ch = os.path.join(work_dir, "segments_4ch")

    output_1ch = os.path.join(aligned_dir, f"{concat_stem}_1ch.wav")
    output_4ch = os.path.join(aligned_dir, f"{concat_stem}_4ch.wav")

    logger.info("========== Step 2: 对齐 ==========")
    stage2_align(
        concat_wav,
        recorded_1ch,
        recorded_4ch,
        output_1ch,
        output_4ch,
        search_range_sec=search_range_sec,
        sample_rate=sample_rate,
    )

    logger.info("========== Step 3: 切分 ==========")
    stage3_split(
        output_1ch,
        concat_tn_txt,
        concat_itn_txt,
        segments_dir,
        segment_sec=segment_sec,
        sample_rate=sample_rate,
        aligned_wav_4ch=output_4ch,
        output_dir_4ch=segments_dir_4ch,
    )

    # 切分后文本：base_name 为对齐 1ch 的 stem，即 {concat_stem}_1ch
    base_name_1ch = Path(output_1ch).stem
    if text_type == "itn":
        text_file = os.path.join(segments_dir, f"{base_name_1ch}_text_itn.txt")
    else:
        text_file = os.path.join(segments_dir, f"{base_name_1ch}_text_tn.txt")

    logger.info("========== Step 4a: 单通道切分 ASR 转写与 WER ==========")
    asr_shared = stage4_asr_eval(
        segments_dir,
        text_file,
        output_excel,
        asr_model=asr_model,
        batch_size=batch_size,
        device=device,
        multichannel_downmix=False,
        preloaded_asr_model=None,
    )

    base_name_4ch = Path(output_4ch).stem
    if text_type == "itn":
        text_file_4ch = os.path.join(segments_dir_4ch, f"{base_name_4ch}_text_itn.txt")
    else:
        text_file_4ch = os.path.join(segments_dir_4ch, f"{base_name_4ch}_text_tn.txt")

    logger.info("========== Step 4b: 四通道切分 ASR 转写与 WER（多声道下混）==========")
    stage4_asr_eval(
        segments_dir_4ch,
        text_file_4ch,
        output_excel_4ch,
        asr_model=asr_model,
        batch_size=batch_size,
        device=device,
        multichannel_downmix=True,
        preloaded_asr_model=asr_shared,
    )
    logger.info(
        "合并流程完成: 对齐 -> 切分 -> 转写(1ch) -> %s ; 转写(4ch) -> %s",
        output_excel,
        output_excel_4ch,
    )


def parse_args():
    parser = argparse.ArgumentParser(description="Audio concat and split tool")
    sub = parser.add_subparsers(dest="stage")

    p1 = sub.add_parser("concat", help="Stage 1: group by duration and concat")
    p1.add_argument("--wav_scp", required=True)
    p1.add_argument("--text_tn", required=True, help="text_tn 文件路径")
    p1.add_argument("--text_itn", required=True, help="text_itn 文件路径")
    p1.add_argument("--wav2dur", required=True)
    p1.add_argument("--output_dir", required=True)
    p1.add_argument("--sr", type=int, default=24000)

    p2 = sub.add_parser("align", help="Stage 2: align recorded PCM audio (1ch + 4ch)")
    p2.add_argument("--concat_wav", required=True, help="原始拼接WAV (参考音频)")
    p2.add_argument("--recorded_1ch", required=True, help="录制的单通道16k PCM文件")
    p2.add_argument("--recorded_4ch", required=True, help="录制的4通道16k PCM文件")
    p2.add_argument("--output_1ch", required=True, help="对齐后的单通道WAV输出路径")
    p2.add_argument("--output_4ch", required=True, help="对齐后的4通道WAV输出路径")
    p2.add_argument("--search_range", type=float, default=300.0)
    p2.add_argument("--sr", type=int, default=SAMPLE_RATE)

    p3 = sub.add_parser("split", help="Stage 3: split aligned audio (1ch + optional 4ch)")
    p3.add_argument("--aligned_wav", required=True, help="对齐后的单通道WAV")
    p3.add_argument("--concat_tn_txt", required=True, help="Phase1 输出的 *_tn.txt")
    p3.add_argument("--concat_itn_txt", required=True, help="Phase1 输出的 *_itn.txt")
    p3.add_argument("--output_dir", required=True, help="单通道切分结果输出目录")
    p3.add_argument("--segment_sec", type=float, required=True)
    p3.add_argument("--sr", type=int, default=SAMPLE_RATE)
    p3.add_argument("--aligned_wav_4ch", default=None, help="对齐后的4通道WAV，指定则同时切分4通道")
    p3.add_argument("--output_dir_4ch", default=None, help="4通道切分结果输出目录（与 --aligned_wav_4ch 成对使用）")

    p4 = sub.add_parser("asr_eval", help="Stage 4: 使用 Qwen3-ASR-1.7B 识别切分音频并算 WER，输出 Excel")
    p4.add_argument("--segments_dir", required=True, help="Stage3 切分结果目录（含 wav 与对应 text 文件所在目录）")
    p4.add_argument("--text_file", required=True, help="切分文本列表，格式: segment_id\\ttext（如 *_text_tn.txt）")
    p4.add_argument("--output_excel", required=True, help="输出 Excel 路径（.xlsx），列: wav_name, text, asr, wer")
    p4.add_argument("--asr_model", default="./Qwen/Qwen3-ASR-1.7B", help="ASR 模型名称，默认 Qwen3-ASR-1.7B")
    p4.add_argument("--batch_size", type=int, default=1, help="Qwen3-ASR 批推理大小，默认 16")
    p4.add_argument("--device", default="cuda:0", help="推理使用的 GPU，如 cuda:0 / cuda:1，避免多卡时张量设备不一致")
    p4.add_argument(
        "--multichannel_downmix",
        action="store_true",
        help="多声道 WAV 按通道平均下混为单声道再识别（4 通道切分片段需加此选项）",
    )

    p_merge = sub.add_parser(
        "align_split_asr",
        help="合并执行: 对齐(2) -> 切分(3) -> ASR转写(4)，中间结果落在 --work_dir",
    )
    p_merge.add_argument(
        "--input_dir",
        default=None,
        help="输入目录：内含 1 个 .wav、各 1 个以 ch1.pcm / ch4.pcm / _tn.txt / _itn.txt 结尾的文件；"
        "若指定则不必再传 --concat_wav 等五个路径",
    )
    p_merge.add_argument("--concat_wav", default=None, help="原始拼接 WAV（参考音频）；与 --input_dir 二选一")
    p_merge.add_argument("--recorded_1ch", default=None, help="录制的单通道 16k PCM")
    p_merge.add_argument("--recorded_4ch", default=None, help="录制的 4 通道 16k PCM")
    p_merge.add_argument("--concat_tn_txt", default=None, help="Stage1 输出的 *_tn.txt")
    p_merge.add_argument("--concat_itn_txt", default=None, help="Stage1 输出的 *_itn.txt")
    p_merge.add_argument("--segment_sec", type=float, required=True, help="每段时长（秒），与 Stage1 分组一致，如 10")
    p_merge.add_argument("--work_dir", required=True, help="工作目录：aligned/、segments/、segments_4ch/ 将在此下生成")
    p_merge.add_argument(
        "--output_excel",
        default=None,
        help="单通道 WER Excel；默认 work_dir/asr_wer_ch1.xlsx",
    )
    p_merge.add_argument(
        "--output_excel_4ch",
        default=None,
        help="四通道 WER Excel；默认 work_dir/asr_wer_ch4.xlsx",
    )
    p_merge.add_argument("--search_range", type=float, default=300.0, help="对齐搜索范围（秒）")
    p_merge.add_argument("--sr", type=int, default=SAMPLE_RATE)
    p_merge.add_argument("--text_type", choices=("tn", "itn"), default="tn", help="ASR 用哪类文本算 WER，默认 tn")
    p_merge.add_argument("--asr_model", default="./Qwen/Qwen3-ASR-1.7B")
    p_merge.add_argument("--batch_size", type=int, default=1)
    p_merge.add_argument("--device", default="cuda:0")

    p5 = sub.add_parser(
        "filter_wer",
        help="Stage 5: 仅按单通道 WER 筛选，四通道按同行下标同步；结果写入 segments/ 与 segments_4ch/；完成后删除 aligned/",
    )
    p5.add_argument(
        "--output_dir",
        required=True,
        help="第四步工作目录（读 asr_wer_ch1.xlsx 做筛选；读 asr_wer_ch4.xlsx 仅按 ch1 保留行同步）",
    )
    p5.add_argument(
        "--global_max_wer",
        type=float,
        default=1.0,
        help="仅作用于单通道：丢弃 wer 大于该值的行；默认 1.0",
    )
    p5.add_argument(
        "--ranges",
        nargs="+",
        required=True,
        metavar="START:END:THR",
        help="仅作用于单通道 Excel：闭区间 [START,END] 内仅保留 wer<=THR（并与全局上限取更严者），"
        "数据行 0 起始下标。例: 400:700:0.45 774:903:0.15；四通道取相同行下标，不再按 4ch 的 wer 筛选",
    )
    p5.add_argument(
        "--excel_ch1",
        default=None,
        help="单通道 Excel 路径；默认 output_dir/asr_wer_ch1.xlsx",
    )
    p5.add_argument(
        "--excel_ch4",
        default=None,
        help="多通道 Excel 路径；默认 output_dir/asr_wer_ch4.xlsx",
    )

    return parser.parse_args()


def main():
    args = parse_args()
    normalize_cli_paths(args)
    if args.stage == "concat":
        stage1_concat(args.wav_scp, args.text_tn, args.text_itn,
                      args.wav2dur, args.output_dir, args.sr)
    elif args.stage == "align":
        stage2_align(args.concat_wav, args.recorded_1ch, args.recorded_4ch,
                     args.output_1ch, args.output_4ch,
                     args.search_range, args.sr)
    elif args.stage == "split":
        stage3_split(
            args.aligned_wav,
            args.concat_tn_txt,
            args.concat_itn_txt,
            args.output_dir,
            args.segment_sec,
            args.sr,
            aligned_wav_4ch=args.aligned_wav_4ch,
            output_dir_4ch=args.output_dir_4ch,
        )
    elif args.stage == "asr_eval":
        stage4_asr_eval(
            args.segments_dir,
            args.text_file,
            args.output_excel,
            asr_model=args.asr_model,
            batch_size=args.batch_size,
            device=args.device,
            multichannel_downmix=args.multichannel_downmix,
        )
    elif args.stage == "align_split_asr":
        if getattr(args, "input_dir", None):
            try:
                resolved = resolve_align_split_input_dir(args.input_dir)
            except ValueError as e:
                logger.error("%s", e)
                sys.exit(1)
            cw = resolved["concat_wav"]
            r1 = resolved["recorded_1ch"]
            r4 = resolved["recorded_4ch"]
            tn = resolved["concat_tn_txt"]
            itn = resolved["concat_itn_txt"]
        else:
            if not all(
                (
                    args.concat_wav,
                    args.recorded_1ch,
                    args.recorded_4ch,
                    args.concat_tn_txt,
                    args.concat_itn_txt,
                )
            ):
                logger.error(
                    "align_split_asr 请提供 --input_dir，或同时提供 "
                    "--concat_wav --recorded_1ch --recorded_4ch --concat_tn_txt --concat_itn_txt"
                )
                sys.exit(1)
            cw, r1, r4, tn, itn = (
                args.concat_wav,
                args.recorded_1ch,
                args.recorded_4ch,
                args.concat_tn_txt,
                args.concat_itn_txt,
            )
        run_align_split_asr(
            cw,
            r1,
            r4,
            tn,
            itn,
            args.segment_sec,
            args.work_dir,
            output_excel=args.output_excel,
            output_excel_4ch=args.output_excel_4ch,
            search_range_sec=args.search_range,
            sample_rate=args.sr,
            text_type=args.text_type,
            asr_model=args.asr_model,
            batch_size=args.batch_size,
            device=args.device,
        )
    elif args.stage == "filter_wer":
        try:
            stage5_filter_wer(
                args.output_dir,
                list(args.ranges),
                global_max_wer=args.global_max_wer,
                excel_ch1=args.excel_ch1,
                excel_ch4=args.excel_ch4,
            )
        except ValueError as e:
            logger.error("%s", e)
            sys.exit(1)
    else:
        logger.error(
            "please specify stage: concat / align / split / asr_eval / align_split_asr / filter_wer"
        )


if __name__ == "__main__":
#     # Phase 1
# python run_audio_cat_cut.py concat \
#     --wav_scp data/wav.scp \
#     --text_tn data/text_tn \
#     --text_itn data/text_itn \
#     --wav2dur data/wav2dur \
#     --output_dir output/concat

# # Phase 2（1ch + 4ch PCM 对齐）
# python run_audio_cat_cut.py align \
#     --concat_wav output/concat/10s_01.wav \
#     --recorded_1ch recorded/10s_01_ch1.pcm \
#     --recorded_4ch recorded/10s_01_ch1.pcm \
#     --output_1ch output/aligned/10s_01_ch1.wav \
#     --output_4ch output/aligned/10s_01_ch4.wav

# # Phase 3（1ch + 4ch 按拼接时长切分）
# python run_audio_cat_cut.py split \
#     --aligned_wav output/aligned/10s_01_ch1.wav \
#     --concat_tn_txt output/concat/10s_01_tn.txt \
#     --concat_itn_txt output/concat/10s_01_itn.txt \
#     --output_dir output/segments_ch1 \
#     --segment_sec 10 \
#     --aligned_wav_4ch output/aligned/10s_01_ch4.wav \
#     --output_dir_4ch output/segments_ch4
#
# # Phase 4（切分音频 ASR 识别 + WER，输出 Excel）
# python run_audio_cat_cut.py asr_eval \
#     --segments_dir output/segments_ch1 \
#     --text_file output/segments/10s_01_ch1_text_tn.txt \
#     --output_excel output/asr_wer_10s_01_ch1.xlsx
#
# # 合并：对齐 + 切分 + 转写（一步跑完 2/3/4）
# python run_audio_cat_cut.py align_split_asr \
#     --concat_wav output/concat/10s_01.wav \
#     --recorded_1ch output/concat/10s_01_ch1.pcm \
#     --recorded_4ch output/concat/10s_01_ch4.pcm \
#     --concat_tn_txt output/concat/10s_01_tn.txt \
#     --concat_itn_txt output/concat/10s_01_itn.txt \
#     --segment_sec 10 \
#     --work_dir output/pipeline_10s_01
#     # 省略 --output_excel / --output_excel_4ch 时默认为 work_dir/asr_wer_ch1.xlsx 与 work_dir/asr_wer_ch4.xlsx
#
# # 合并（使用 input_dir：目录内需 1 个 .wav、各 1 个 ch1.pcm / ch4.pcm / _tn.txt / _itn.txt）
# python run_audio_cat_cut.py align_split_asr \
#     --input_dir /path/to/batch_inputs \
#     --segment_sec 10 \
#     --work_dir /path/to/output_work
#
# # Phase 5（仅单通道按 WER/区间筛选，4ch 按同行下标同步；数据行下标从 0 起）
# python run_audio_cat_cut.py filter_wer \
#     --output_dir /path/to/output_work \
#     --global_max_wer 1.0 \
#     --ranges 400:700:0.45 774:903:0.15

    main()